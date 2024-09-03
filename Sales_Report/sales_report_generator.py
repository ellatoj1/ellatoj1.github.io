import openpyxl  
import csv  
import datetime  
import os  

from openpyxl.chart import BarChart, Series, Reference  

dir_path = os.path.dirname(__file__)
csv_path = os.path.join(dir_path, 'sales_data.csv')  
output_path = os.path.join(os.path.dirname(__file__), 'output.xlsx')  

SALES_SHEET_NAME = 'Sales_sheet'
REGION_SHEET_NAME = 'Region_sheet'

# create a new Excel workbook and set the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = SALES_SHEET_NAME

sales_sheet = workbook[SALES_SHEET_NAME]
region_sheet = workbook.create_sheet(REGION_SHEET_NAME)

with open(csv_path) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    data = list(csv_reader)

# set headers in the first row of the Sales sheet
headers = data[0]
for column_index, header in enumerate(headers):
    worksheet.cell(row=1, column=column_index+1, value=str(header))

# sort the data by date and write it to the Excel sheet
data = sorted(data[1:], key=lambda x: datetime.datetime.strptime(x[0], "%Y-%m-%d"))
for row_index, row_data in enumerate(data, start=2):
    for column_index, cell_data in enumerate(row_data):
        worksheet.cell(row=row_index, column=column_index+1, value=cell_data)

# add a column for total sales
worksheet.cell(row=1, column=6, value='Total försäljning')
for row_index in range(2, worksheet.max_row + 1):
    count = int(worksheet.cell(row=row_index, column=4).value)
    price = float(worksheet.cell(row=row_index, column=5).value)
    total_amount = count * price
    worksheet.cell(row=row_index, column=6, value=total_amount).number_format = "#,##0"

for cell in worksheet[1]:
    cell.font = openpyxl.styles.Font(bold=True, size=14)



""" REGION SALES TABLE ON A SEPARATE SHEET """

# calculate sales per region and store in a dictionary
region_sales = {}
for row_index in range(2, worksheet.max_row + 1):
    region = worksheet.cell(row=row_index, column=3).value
    total_sales = float(worksheet.cell(row=row_index, column=6).value)
    region_sales[region] = round(region_sales.get(region, 0) + total_sales, 2)

# switch to the region sales sheet
worksheet = workbook[region_sheet.title]

# set headers for the region sales sheet
worksheet.cell(row=1, column=1, value="Region")
worksheet.cell(row=1, column=2, value="Sales")

for cell in worksheet[1]:
    cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF', size=14)

table_range = f"A1:B{len(region_sales) + 1}"

table = openpyxl.worksheet.table.Table(displayName="Region_Sales", ref=table_range)

style = openpyxl.worksheet.table.TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
table.tableStyleInfo = style

worksheet.add_table(table)

# write the sales data per region to the region sheet
for index, (name, sales) in enumerate(sorted(region_sales.items()), start=2):
    worksheet.cell(row=index, column=1, value=name)
    worksheet.cell(row=index, column=2, value=sales).number_format = "#,##0"



""" CREATE A BAR CHART """

data = openpyxl.chart.Reference(worksheet, min_col=2, min_row=1, max_col=2, max_row=worksheet.max_row)

categories = openpyxl.chart.Reference(worksheet, min_col=1, min_row=2, max_col=1, max_row=worksheet.max_row)

chart = openpyxl.chart.BarChart()
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Försäljning per region"
chart.x_axis.title = "Region"
chart.y_axis.title = "Försäljning (kr)"

worksheet.add_chart(chart, "D2")

workbook.save(output_path)
